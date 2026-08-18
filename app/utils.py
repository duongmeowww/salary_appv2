import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REQUIRED_COLUMNS = (
    'Mã nhân viên',
    'Họ tên',
    'Tháng',
    'Năm',
    'Lương cơ bản',
    'Phụ cấp',
    'Khấu trừ',
    'Lương thực nhận',
)

OPTIONAL_COLUMNS = {
    'department': ('Phòng ban', 'Bộ phận', 'Xưởng', 'Chuyền may'),
    'position': ('Chức vụ', 'Vị trí'),
    'phone': ('Số điện thoại', 'SĐT', 'SDT', 'Điện thoại'),
    'bank_name': ('Ngân hàng', 'Tên ngân hàng'),
    'bank_account': ('Số tài khoản', 'STK', 'Số TK'),
}


def _to_float(value, column, row_number):
    if value is None or value == '':
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f'Dòng {row_number}: cột "{column}" phải là số, nhận được "{value}".')


def _to_int(value, column, row_number, minimum, maximum):
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f'Dòng {row_number}: cột "{column}" phải là số nguyên, nhận được "{value}".')
    if not minimum <= number <= maximum:
        raise ValueError(f'Dòng {row_number}: cột "{column}" phải nằm trong khoảng {minimum}-{maximum}.')
    return number


def parse_salary_excel(source):
    """Đọc bảng lương từ file Excel (đường dẫn hoặc file-like object).

    Dùng openpyxl ở chế độ read-only nên chỉ giữ từng dòng trong bộ nhớ khi duyệt.
    """
    if hasattr(source, 'read') and not hasattr(source, 'seekable'):
        # SpooledTemporaryFile (upload của Werkzeug trên Python < 3.11) không có seekable()
        source = io.BytesIO(source.read())

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError('File Excel rỗng.')

        indexes = {}
        for position, name in enumerate(header):
            if isinstance(name, str):
                indexes.setdefault(name.strip(), position)

        missing = [column for column in REQUIRED_COLUMNS if column not in indexes]
        if missing:
            raise ValueError(f"File Excel thiếu cột bắt buộc: {', '.join(missing)}")

        records = []
        for row_number, row in enumerate(rows, start=2):
            def cell(column):
                position = indexes[column]
                return row[position] if position < len(row) else None

            username = cell('Mã nhân viên')
            if username is None or str(username).strip() == '':
                continue

            def optional_cell(field_key):
                aliases = OPTIONAL_COLUMNS.get(field_key, ())
                for alias in aliases:
                    if alias in indexes:
                        pos = indexes[alias]
                        val = row[pos] if pos < len(row) else None
                        if val is not None and str(val).strip() != '':
                            return str(val).strip()
                return None

            records.append({
                'username': str(username).strip(),
                'full_name': str(cell('Họ tên') or '').strip(),
                'month': _to_int(cell('Tháng'), 'Tháng', row_number, 1, 12),
                'year': _to_int(cell('Năm'), 'Năm', row_number, 1900, 2999),
                'basic_salary': _to_float(cell('Lương cơ bản'), 'Lương cơ bản', row_number),
                'allowance': _to_float(cell('Phụ cấp'), 'Phụ cấp', row_number),
                'deduction': _to_float(cell('Khấu trừ'), 'Khấu trừ', row_number),
                'net_salary': _to_float(cell('Lương thực nhận'), 'Lương thực nhận', row_number),
                'department': optional_cell('department'),
                'position': optional_cell('position'),
                'phone': optional_cell('phone'),
                'bank_name': optional_cell('bank_name'),
                'bank_account': optional_cell('bank_account'),
            })
        return records
    finally:
        workbook.close()


def generate_sample_excel_stream():
    """Tạo file Excel mẫu chuẩn trả về dưới dạng BytesIO stream để tải về từ Web."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'BangLuongMau'

    headers = [
        'Mã nhân viên',
        'Họ tên',
        'Tháng',
        'Năm',
        'Lương cơ bản',
        'Phụ cấp',
        'Khấu trừ',
        'Lương thực nhận',
        'Phòng ban',
        'Chức vụ',
        'Số điện thoại',
    ]
    ws.append(headers)

    sample_data = [
        ['NV001', 'Nguyễn Văn An', 10, 2025, 8500000, 1200000, 500000, 9200000, 'Chuyền may 1', 'Công nhân may', '0912345678'],
        ['NV002', 'Trần Thị Bình', 10, 2025, 9000000, 1500000, 600000, 9900000, 'Chuyền may 2', 'Tổ phó chuyền', '0987654321'],
        ['NV003', 'Lê Hoàng Cường', 10, 2025, 10500000, 2000000, 800000, 11700000, 'Tổ Cắt', 'Thợ cắt chính', '0903112233'],
        ['NV004', 'Phạm Thị Dung', 10, 2025, 8000000, 1000000, 400000, 8600000, 'Tổ Hoàn thiện', 'Ủi đóng gói', '0934556677'],
        ['NV005', 'Vũ Minh Đức', 10, 2025, 11000000, 2500000, 1000000, 12500000, 'Phòng Kỹ thuật', 'Kỹ thuật may', '0945667788'],
    ]

    for row in sample_data:
        ws.append(row)

    header_fill = PatternFill(start_color='10294A', end_color='10294A', fill_type='solid')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE'),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    data_font = Font(name='Arial', size=10)
    for row in ws.iter_rows(min_row=2, max_row=len(sample_data) + 1):
        ws.row_dimensions[row[0].row].height = 22
        for col_idx, cell in enumerate(row):
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (0, 2, 3):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in (1, 8, 9, 10):
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '#,##0'

    col_widths = {
        'A': 15, 'B': 22, 'C': 10, 'D': 10, 'E': 15,
        'F': 14, 'G': 14, 'H': 18, 'I': 18, 'J': 18, 'K': 16
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
