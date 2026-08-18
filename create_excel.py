import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Tạo workbook
wb = Workbook()
ws = wb.active
ws.title = 'BangLuongThang10_2025'

# Tiêu đề cột chuẩn theo hệ thống
headers = [
    'Mã nhân viên',
    'Họ tên',
    'Tháng',
    'Năm',
    'Lương cơ bản',
    'Phụ cấp',
    'Khấu trừ',
    'Lương thực nhận',
]
ws.append(headers)

# Dữ liệu mẫu nhân viên xưởng may
data = [
    ['NV001', 'Nguyễn Văn An', 10, 2025, 8500000, 1200000, 500000, 9200000],
    ['NV002', 'Trần Thị Bình', 10, 2025, 9000000, 1500000, 600000, 9900000],
    ['NV003', 'Lê Hoàng Cường', 10, 2025, 10500000, 2000000, 800000, 11700000],
    ['NV004', 'Phạm Thị Dung', 10, 2025, 8000000, 1000000, 400000, 8600000],
    ['NV005', 'Vũ Minh Đức', 10, 2025, 11000000, 2500000, 1000000, 12500000],
    ['NV006', 'Đỗ Thị Hạnh', 10, 2025, 8200000, 1100000, 450000, 8850000],
    ['NV007', 'Ngô Văn Hùng', 10, 2025, 9500000, 1800000, 700000, 10600000],
    ['NV008', 'Hoàng Thị Lan', 10, 2025, 8800000, 1300000, 550000, 9550000],
    ['NV009', 'Đặng Quốc Nam', 10, 2025, 12000000, 3000000, 1200000, 13800000],
    ['NV010', 'Bùi Thị Phương', 10, 2025, 8600000, 1250000, 520000, 9330000],
    ['NV011', 'Phan Thanh Sơn', 10, 2025, 9200000, 1600000, 650000, 10150000],
    ['NV012', 'Lý Thu Thảo', 10, 2025, 8400000, 1150000, 480000, 9070000],
]

for row in data:
    ws.append(row)

# Định dạng header
header_fill = PatternFill(start_color='10294A', end_color='10294A', fill_type='solid')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin', color='D0D7DE'),
    right=Side(style='thin', color='D0D7DE'),
    top=Side(style='thin', color='D0D7DE'),
    bottom=Side(style='thin', color='D0D7DE'),
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[1].height = 28

# Định dạng dữ liệu các dòng
data_font = Font(name='Arial', size=10)
for row in ws.iter_rows(min_row=2, max_row=len(data) + 1):
    ws.row_dimensions[row[0].row].height = 22
    for col_idx, cell in enumerate(row):
        cell.font = data_font
        cell.border = thin_border
        if col_idx in (0, 2, 3):  # Mã NV, Tháng, Năm
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 1:  # Họ tên
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:  # Các cột tiền
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0'

# Căn chỉnh độ rộng cột tự động
col_widths = {
    'A': 16,  # Mã nhân viên
    'B': 22,  # Họ tên
    'C': 10,  # Tháng
    'D': 10,  # Năm
    'E': 16,  # Lương cơ bản
    'F': 15,  # Phụ cấp
    'G': 15,  # Khấu trừ
    'H': 18,  # Lương thực nhận
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Lưu file
output_path = 'bang_luong_mau.xlsx'
wb.save(output_path)
print(f'File {output_path} da duoc tao thanh cong!')
